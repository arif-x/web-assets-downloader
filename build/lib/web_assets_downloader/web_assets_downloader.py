import os
from urllib.parse import urlparse
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, unquote
from PIL import Image
import PyPDF2
from docx import Document
from openpyxl import load_workbook

def download_asset(asset_urls, save_folder, headers):
    for url in asset_urls:
        print(url)
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            # Parse the asset URL to extract the path
            parsed_url = urlparse(url)
            asset_path = unquote(parsed_url.path)

            # Construct the filepath within the assets directory
            asset_filepath = os.path.join(save_folder, asset_path.lstrip('/'))

            # Create the directory if it doesn't exist
            os.makedirs(os.path.dirname(asset_filepath), exist_ok=True)

            # Download the asset
            with open(asset_filepath, 'wb') as f:
                f.write(response.content)
                print(f"Downloaded {asset_path}.")

            # Resize images if it's an image file
            if asset_path.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                resize_image(asset_filepath, (100, 100))  # Specify desired dimensions

            # Extract text from PDF files
            if asset_path.lower().endswith('.pdf'):
                extract_text_from_pdf(asset_filepath)

            # Extract text from Word files
            if asset_path.lower().endswith('.docx'):
                extract_text_from_docx(asset_filepath)

            # Extract text from Excel files
            if asset_path.lower().endswith('.xlsx'):
                extract_text_from_excel(asset_filepath)

def resize_image(image_path, dimensions):
    img = Image.open(image_path)
    img_resized = img.resize(dimensions, Image.ANTIALIAS)
    img_resized.save(image_path)

def extract_text_from_pdf(pdf_path):
    with open(pdf_path, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfFileReader(pdf_file)
        text = ''
        for page_num in range(pdf_reader.numPages):
            page = pdf_reader.getPage(page_num)
            text += page.extractText()
        print(f"Text extracted from {pdf_path}: {text}")

def extract_text_from_docx(docx_path):
    doc = Document(docx_path)
    text = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
    print(f"Text extracted from {docx_path}: {text}")

def extract_text_from_excel(excel_path):
    wb = load_workbook(excel_path)
    text = ''
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                text += str(cell.value) + ' '
            text += '\n'
    print(f"Text extracted from {excel_path}: {text}")

def download_html_and_asset(url_list, save_folder, max_depth=None):
    # Standard User-Agent header for a common web browser
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

    for url in url_list:
        response = requests.get(url, headers=headers)
        print(response.status_code)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            base_url = f"{urlparse(url).scheme}://{urlparse(url).netloc}"

            # Extract the filename from the last part of the URL path
            filename = os.path.basename(urlparse(url).path)
            if not filename:
                filename = "index.html"

            # Save HTML content
            html_filepath = os.path.join(save_folder, filename)

            # Create the directory if it doesn't exist
            os.makedirs(os.path.dirname(html_filepath), exist_ok=True)

            with open(html_filepath, 'wb') as f:
                f.write(response.content)
                print(f"HTML content saved for {url} as {filename}.")

            # Extract asset URLs
            asset_urls = set()
            for tag in soup.find_all(['img', 'link', 'script', 'a']):
                if tag.name == 'a' and max_depth is not None:
                    # Limit recursion depth
                    if max_depth <= 0:
                        continue
                    max_depth -= 1
                    asset_url = tag['href']
                elif tag.has_attr('src'):
                    asset_url = tag['src']
                elif tag.has_attr('href'):
                    asset_url = tag['href']
                else:
                    continue

                absolute_url = urljoin(base_url, asset_url)

                print(absolute_url)
                asset_urls.add(absolute_url)

            # Download assets
            download_asset(asset_urls, save_folder, headers)

    return True
